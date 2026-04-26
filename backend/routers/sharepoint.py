"""
SharePoint / OneDrive Integration Router (SharePoint 2.0)
─────────────────────────────────────────────────────────
Implements the full Microsoft OAuth2 flow with optional tenant scoping and
STREAMING downloads so large files (100MB+ Excel workbooks) never balloon
the backend's memory footprint.

Endpoints:
  1. GET    /auth-url                  → returns MS login URL (tenant-scoped if tenant provided)
  2. GET    /callback                  → exchanges code for tokens, stores them, redirects
  3. GET    /status                    → is the current org connected?
  4. DELETE /disconnect                → revoke & delete stored tokens
  5. GET    /drives                    → list accessible OneDrive + SharePoint drives
  6. GET    /files                     → list files/folders inside a drive or folder
  7. POST   /import                    → STREAM a file from SP/OD and import it (body supports `keep_linked`)
  8. GET    /linked-files              → list every linked SP/OD file in the org
  9. POST   /{file_id}/resync          → re-stream the latest bytes for a linked file
 10. POST   /{file_id}/sync-schedule   → off / hourly / daily auto-resync

Tenant scoping:
  - Pass ?tenant=contoso.com (domain) or ?tenant=<uuid> to /auth-url
  - Without tenant → uses /common/ endpoint (personal + any org accounts)
  - With tenant    → uses /{tenant}/ endpoint (locks to that org's Azure AD)

Required env vars (set in Railway service):
  MICROSOFT_CLIENT_ID       – Azure AD app client ID
  MICROSOFT_CLIENT_SECRET   – Azure AD app client secret
  MICROSOFT_REDIRECT_URI    – e.g. https://<backend>.up.railway.app/api/sharepoint/callback

Streaming strategy:
  - Microsoft Graph returns a redirect to a pre-signed URL, followed by the
    actual file bytes. httpx handles redirects for us.
  - We pipe bytes into a SpooledTemporaryFile (RAM below 8MB, disk above).
  - For R2/S3 we hand the spooled file to boto3 upload_fileobj (chunked multi-part).
  - For local storage we rename the temp file into place (no re-read).
  - openpyxl/csv parse the spooled file once — the bytes are already on disk
    so we don't pull them back into RAM.
"""
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import RedirectResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel, Field
from typing import Optional
from datetime import datetime, timedelta
import httpx
import uuid
import json
import tempfile
import os
import shutil

from database import get_db, User, DataFile
from auth_utils import get_current_user
from config import settings

router = APIRouter()

GRAPH_BASE    = "https://graph.microsoft.com/v1.0"
SCOPES        = "Files.Read Sites.Read.All offline_access User.Read"
ALLOWED_EXT   = {".xlsx", ".xls", ".csv"}
# Max file size we'll accept. Goes to disk, not RAM, but still a safety valve.
MAX_IMPORT_BYTES = 500 * 1024 * 1024  # 500 MB
STREAM_CHUNK     = 512 * 1024         # 512 KB per read
SPOOL_MAX_SIZE   = 8 * 1024 * 1024    # 8 MB kept in RAM; above → /tmp

VALID_SYNC_FREQUENCIES = {"off", "hourly", "daily"}


def _oauth_base(tenant: str | None = None) -> str:
    """Return the correct Microsoft OAuth base URL for the given tenant.
    Falls back to /common/ if no tenant specified (personal + any org)."""
    scope = tenant if tenant else "common"
    return f"https://login.microsoftonline.com/{scope}/oauth2/v2.0"


# ── Token persistence helpers ─────────────────────────────────────────────────

def _get_org_token(db: Session, org_id: str):
    return db.execute(
        text(
            "SELECT access_token, refresh_token, expires_at, tenant_id "
            "FROM sharepoint_tokens WHERE organisation_id = :oid"
        ),
        {"oid": org_id},
    ).fetchone()


def _save_org_token(
    db: Session, org_id: str, access_token: str,
    refresh_token: str, expires_in: int, tenant_id: str | None = None,
):
    expires_at = (datetime.utcnow() + timedelta(seconds=expires_in)).isoformat()
    db.execute(
        text("""
            INSERT INTO sharepoint_tokens
                (id, organisation_id, access_token, refresh_token, expires_at, tenant_id, created_at)
            VALUES (:id, :oid, :at, :rt, :ea, :tid, NOW())
            ON CONFLICT (organisation_id) DO UPDATE
               SET access_token  = EXCLUDED.access_token,
                   refresh_token = EXCLUDED.refresh_token,
                   expires_at    = EXCLUDED.expires_at,
                   tenant_id     = EXCLUDED.tenant_id
        """),
        {
            "id":  str(uuid.uuid4()),
            "oid": org_id,
            "at":  access_token,
            "rt":  refresh_token,
            "ea":  expires_at,
            "tid": tenant_id,
        },
    )
    db.commit()


async def _get_valid_access_token(db: Session, org_id: str) -> str:
    """Return a fresh access token, refreshing via refresh_token if needed."""
    row = _get_org_token(db, org_id)
    if not row:
        raise HTTPException(
            status_code=401,
            detail="SharePoint not connected. Connect via Integrations → SharePoint.",
        )
    access_token, refresh_token, expires_at_str, tenant_id = row
    try:
        expires_at = datetime.fromisoformat(expires_at_str)
    except Exception:
        expires_at = datetime.utcnow() - timedelta(minutes=1)

    if datetime.utcnow() > expires_at - timedelta(minutes=5):
        oauth_base = _oauth_base(tenant_id)
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{oauth_base}/token",
                data={
                    "client_id":     settings.MICROSOFT_CLIENT_ID,
                    "client_secret": settings.MICROSOFT_CLIENT_SECRET,
                    "grant_type":    "refresh_token",
                    "refresh_token": refresh_token,
                    "scope":         SCOPES,
                },
                timeout=20,
            )
        if resp.status_code != 200:
            db.execute(
                text("DELETE FROM sharepoint_tokens WHERE organisation_id = :oid"),
                {"oid": org_id},
            )
            db.commit()
            raise HTTPException(
                status_code=401,
                detail="SharePoint session expired. Please reconnect.",
            )
        data = resp.json()
        new_refresh = data.get("refresh_token", refresh_token)
        _save_org_token(db, org_id, data["access_token"], new_refresh, data["expires_in"], tenant_id)
        return data["access_token"]

    return access_token


# ── Streaming download core ──────────────────────────────────────────────────

async def _stream_graph_file_to_temp(access_token: str, drive_id: str, item_id: str):
    """Stream a Graph file into a SpooledTemporaryFile.
    Returns (spooled_file, total_bytes).
    Raises HTTPException on download failure or size-limit breach."""
    url = f"{GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content"
    tmp = tempfile.SpooledTemporaryFile(max_size=SPOOL_MAX_SIZE)
    total = 0
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=None) as client:
            async with client.stream(
                "GET", url,
                headers={"Authorization": f"Bearer {access_token}"},
            ) as resp:
                if resp.status_code != 200:
                    # Drain and raise — don't swallow the error body blindly
                    body = await resp.aread()
                    raise HTTPException(
                        status_code=resp.status_code,
                        detail=f"Failed to download from SharePoint: {body[:200].decode(errors='ignore')}",
                    )
                async for chunk in resp.aiter_bytes(STREAM_CHUNK):
                    if not chunk:
                        continue
                    total += len(chunk)
                    if total > MAX_IMPORT_BYTES:
                        raise HTTPException(
                            status_code=413,
                            detail=f"File exceeds the {MAX_IMPORT_BYTES // (1024*1024)}MB import limit.",
                        )
                    tmp.write(chunk)
    except HTTPException:
        tmp.close()
        raise
    except Exception as exc:
        tmp.close()
        raise HTTPException(status_code=502, detail=f"Streaming download failed: {exc}")

    tmp.seek(0)
    return tmp, total


def _parse_headers_from_temp(spooled, ext: str) -> tuple[int, int, str]:
    """Parse row_count, column_count, columns_json from a spooled temp file.
    Resets file position to 0 before and after."""
    row_count, column_count, columns_json = 0, 0, "[]"
    spooled.seek(0)
    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(spooled, read_only=True, data_only=True)
            ws = wb.active
            hdr = next(ws.iter_rows(max_row=1, values_only=True), None)
            if hdr:
                headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(hdr)]
                column_count = len(headers)
                columns_json = json.dumps(headers)
            row_count = max(0, (ws.max_row or 1) - 1)
            wb.close()
        elif ext == ".csv":
            import csv, io
            # For very large CSVs we count rows iteratively rather than list()ing
            spooled.seek(0)
            # csv.reader wants text; use utf-8 with replacement for safety
            text_iter = io.TextIOWrapper(spooled, encoding="utf-8", errors="replace", newline="")
            reader = csv.reader(text_iter)
            header_row = next(reader, None)
            if header_row:
                columns_json = json.dumps(header_row)
                column_count = len(header_row)
            row_count = sum(1 for _ in reader)
            # Detach so the underlying spooled file stays open for us to rewind
            text_iter.detach()
    except Exception:
        # Silent — if parsing fails, file is still imported but without row counts
        pass
    spooled.seek(0)
    return row_count, column_count, columns_json


def _persist_temp_file(spooled, filename: str) -> tuple[str, str, bytes | None]:
    """Move the streamed bytes into their final home (R2 or local disk).
    Returns (storage_key, storage_type, file_content_db_or_None).

    We keep `file_content` NULL for both paths — large streamed files should
    never land in Postgres bytes columns. The Files download endpoint already
    handles R2/local fallback for reads.
    """
    spooled.seek(0)
    if settings.STORAGE_TYPE == "s3" and settings.AWS_BUCKET_NAME:
        # Streaming upload — no full-bytes read on the Python side
        from routers.files import _get_s3_client
        s3 = _get_s3_client()
        storage_key = str(uuid.uuid4()) + "_" + filename
        s3.upload_fileobj(spooled, settings.AWS_BUCKET_NAME, storage_key)
        return storage_key, "s3", None

    # Local: stream-copy into LOCAL_UPLOAD_DIR
    os.makedirs(settings.LOCAL_UPLOAD_DIR, exist_ok=True)
    storage_key = str(uuid.uuid4()) + "_" + filename
    dest_path = os.path.join(settings.LOCAL_UPLOAD_DIR, storage_key)
    with open(dest_path, "wb") as dst:
        spooled.seek(0)
        shutil.copyfileobj(spooled, dst, STREAM_CHUNK)
    return storage_key, "local", None


# ── OAuth initiation ──────────────────────────────────────────────────────────

@router.get("/auth-url")
async def get_auth_url(
    tenant: str = Query(None, description="Azure AD tenant domain (e.g. contoso.com) or tenant UUID. Omit for personal/any-org accounts."),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return the Microsoft OAuth URL. Frontend opens this to start the flow.
    Pass ?tenant=contoso.com to lock auth to a specific organisation's Azure AD."""
    if not settings.MICROSOFT_CLIENT_ID:
        raise HTTPException(
            status_code=501,
            detail="SharePoint integration is not configured on this server.",
        )

    # Normalise tenant — strip whitespace and leading https:// if someone pastes a URL
    clean_tenant = None
    if tenant:
        clean_tenant = tenant.strip().lower()
        clean_tenant = clean_tenant.replace("https://", "").replace("http://", "").rstrip("/")
        if "login.microsoftonline.com" in clean_tenant:
            clean_tenant = None  # ignore, treat as common

    state = str(uuid.uuid4())
    expires_at = (datetime.utcnow() + timedelta(minutes=10)).isoformat()
    db.execute(
        text("""
            INSERT INTO oauth_states (state, organisation_id, tenant_id, created_at, expires_at)
            VALUES (:state, :oid, :tid, NOW(), :ea)
        """),
        {"state": state, "oid": current_user.organisation_id, "tid": clean_tenant, "ea": expires_at},
    )
    db.commit()

    oauth_base    = _oauth_base(clean_tenant)
    scope_encoded = SCOPES.replace(" ", "%20")
    redirect_uri  = settings.MICROSOFT_REDIRECT_URI

    url = (
        f"{oauth_base}/authorize"
        f"?client_id={settings.MICROSOFT_CLIENT_ID}"
        f"&response_type=code"
        f"&redirect_uri={redirect_uri}"
        f"&scope={scope_encoded}"
        f"&state={state}"
        f"&prompt=select_account"
    )
    return {"url": url, "tenant": clean_tenant}


@router.get("/callback")
async def oauth_callback(
    code: str  = Query(...),
    state: str = Query(...),
    db: Session = Depends(get_db),
):
    """Exchange auth code for tokens, store them, redirect to frontend."""
    row = db.execute(
        text("SELECT organisation_id, tenant_id, expires_at FROM oauth_states WHERE state = :state"),
        {"state": state},
    ).fetchone()

    frontend_origin = settings.FRONTEND_URL.split(",")[0].strip()

    if not row:
        return RedirectResponse(url=f"{frontend_origin}/sharepoint?error=invalid_state")

    org_id, tenant_id, expires_at_str = row
    if datetime.utcnow() > datetime.fromisoformat(expires_at_str):
        return RedirectResponse(url=f"{frontend_origin}/sharepoint?error=state_expired")

    db.execute(text("DELETE FROM oauth_states WHERE state = :state"), {"state": state})
    db.commit()

    oauth_base = _oauth_base(tenant_id)

    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{oauth_base}/token",
            data={
                "client_id":     settings.MICROSOFT_CLIENT_ID,
                "client_secret": settings.MICROSOFT_CLIENT_SECRET,
                "code":          code,
                "grant_type":    "authorization_code",
                "redirect_uri":  settings.MICROSOFT_REDIRECT_URI,
                "scope":         SCOPES,
            },
            timeout=20,
        )

    if resp.status_code != 200:
        return RedirectResponse(url=f"{frontend_origin}/sharepoint?error=token_failed")

    data = resp.json()
    _save_org_token(
        db, org_id,
        data["access_token"],
        data.get("refresh_token", ""),
        data["expires_in"],
        tenant_id,
    )

    tenant_param = f"&tenant={tenant_id}" if tenant_id else ""
    return RedirectResponse(url=f"{frontend_origin}/sharepoint?connected=true{tenant_param}")


# ── Connection status ─────────────────────────────────────────────────────────

@router.get("/status")
async def get_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    configured = bool(settings.MICROSOFT_CLIENT_ID and settings.MICROSOFT_CLIENT_SECRET)
    row = _get_org_token(db, current_user.organisation_id)
    if not row:
        return {"connected": False, "configured": configured}
    _, _, _, tenant_id = row

    try:
        token = await _get_valid_access_token(db, current_user.organisation_id)
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{GRAPH_BASE}/me",
                headers={"Authorization": f"Bearer {token}"},
                timeout=8,
            )
        if resp.status_code == 200:
            profile = resp.json()
            return {
                "connected":    True,
                "configured":   configured,
                "display_name": profile.get("displayName"),
                "email":        profile.get("mail") or profile.get("userPrincipalName"),
                "tenant":       tenant_id,
            }
    except HTTPException:
        pass

    return {"connected": True, "configured": configured, "display_name": None, "email": None, "tenant": tenant_id}


@router.delete("/disconnect")
async def disconnect(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    db.execute(
        text("DELETE FROM sharepoint_tokens WHERE organisation_id = :oid"),
        {"oid": current_user.organisation_id},
    )
    # Disable auto-resync on all linked SP files — content stays but won't refresh
    db.execute(
        text("""UPDATE data_files SET sync_frequency = 'off'
                WHERE organisation_id = :oid
                  AND sharepoint_item_id IS NOT NULL"""),
        {"oid": current_user.organisation_id},
    )
    db.commit()
    return {"message": "SharePoint disconnected successfully"}


# ── File browser ──────────────────────────────────────────────────────────────

@router.get("/drives")
async def list_drives(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all OneDrive and SharePoint drives accessible to the connected account."""
    token   = await _get_valid_access_token(db, current_user.organisation_id)
    headers = {"Authorization": f"Bearer {token}"}
    drives  = []

    async with httpx.AsyncClient() as client:
        r = await client.get(f"{GRAPH_BASE}/me/drives", headers=headers, timeout=15)
        if r.status_code == 200:
            for d in r.json().get("value", []):
                drives.append({
                    "id":         d["id"],
                    "name":       d.get("name", "My OneDrive"),
                    "type":       "onedrive",
                    "drive_type": d.get("driveType", "personal"),
                })

        r = await client.get(f"{GRAPH_BASE}/me/followedSites", headers=headers, timeout=15)
        if r.status_code == 200:
            for site in r.json().get("value", []):
                site_id = site["id"]
                dr = await client.get(
                    f"{GRAPH_BASE}/sites/{site_id}/drives", headers=headers, timeout=10
                )
                if dr.status_code == 200:
                    for d in dr.json().get("value", []):
                        drives.append({
                            "id":         d["id"],
                            "name":       f"{site.get('displayName', 'SharePoint')} — {d.get('name', 'Documents')}",
                            "type":       "sharepoint",
                            "drive_type": d.get("driveType", "documentLibrary"),
                        })

        r = await client.get(f"{GRAPH_BASE}/sites/root/drives", headers=headers, timeout=10)
        if r.status_code == 200:
            for d in r.json().get("value", []):
                drives.append({
                    "id":         d["id"],
                    "name":       f"Root Site — {d.get('name', 'Documents')}",
                    "type":       "sharepoint",
                    "drive_type": d.get("driveType", "documentLibrary"),
                })

    seen, unique = set(), []
    for d in drives:
        if d["id"] not in seen:
            seen.add(d["id"])
            unique.append(d)
    return {"drives": unique}


@router.get("/files")
async def list_files(
    drive_id:  str = Query(...),
    folder_id: str = Query(None, description="Item ID of the folder; omit for root"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List files and sub-folders inside a drive or folder."""
    token   = await _get_valid_access_token(db, current_user.organisation_id)
    headers = {"Authorization": f"Bearer {token}"}

    url = (
        f"{GRAPH_BASE}/drives/{drive_id}/items/{folder_id}/children"
        if folder_id else
        f"{GRAPH_BASE}/drives/{drive_id}/root/children"
    )

    async with httpx.AsyncClient() as client:
        r = await client.get(
            url, headers=headers,
            params={"$select": "id,name,size,file,folder,lastModifiedDateTime,webUrl"},
            timeout=15,
        )

    if r.status_code != 200:
        raise HTTPException(status_code=r.status_code, detail="Failed to list files from SharePoint")

    items = []
    for item in r.json().get("value", []):
        is_file    = "file" in item
        name       = item["name"]
        ext        = ("." + name.rsplit(".", 1)[-1].lower()) if is_file and "." in name else ""
        importable = ext in ALLOWED_EXT

        items.append({
            "id":         item["id"],
            "name":       name,
            "type":       "file" if is_file else "folder",
            "size":       item.get("size", 0),
            "modified":   item.get("lastModifiedDateTime"),
            "web_url":    item.get("webUrl"),
            "importable": importable,
            "extension":  ext,
        })

    items.sort(key=lambda x: (0 if x["type"] == "folder" else 1, x["name"].lower()))
    return {"items": items}


# ── Import (streaming) ────────────────────────────────────────────────────────

class ImportRequest(BaseModel):
    drive_id: str
    item_id:  str
    filename: str = "imported.xlsx"
    keep_linked: bool = Field(False, description="If true, store drive/item IDs so this file can be re-synced later.")
    sync_frequency: Optional[str] = Field(None, description="Optional: off/hourly/daily. Requires keep_linked=true.")


@router.post("/import")
async def import_file(
    body: ImportRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Stream a file from SharePoint/OneDrive into DataHub Pro.
    Passing keep_linked=true stores the SP identifiers so the file can be
    resynced (either manually via /resync or on an auto-sync schedule)."""
    ext = ("." + body.filename.rsplit(".", 1)[-1].lower()) if "." in body.filename else ""
    if ext not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, and .csv files can be imported.")

    # Validate sync_frequency choice early
    sf = (body.sync_frequency or "off").lower()
    if sf not in VALID_SYNC_FREQUENCIES:
        raise HTTPException(
            status_code=400,
            detail=f"sync_frequency must be one of: {', '.join(sorted(VALID_SYNC_FREQUENCIES))}",
        )
    if sf != "off" and not body.keep_linked:
        raise HTTPException(status_code=400, detail="Auto-sync requires keep_linked=true.")

    token = await _get_valid_access_token(db, current_user.organisation_id)
    spooled, total_bytes = await _stream_graph_file_to_temp(token, body.drive_id, body.item_id)
    try:
        row_count, column_count, columns_json = _parse_headers_from_temp(spooled, ext)
        storage_key, storage_type, _ = _persist_temp_file(spooled, body.filename)
    finally:
        spooled.close()

    file_id = str(uuid.uuid4())
    db_file = DataFile(
        id=file_id,
        filename=storage_key,
        original_filename=body.filename,
        file_size=total_bytes,
        row_count=row_count,
        column_count=column_count,
        columns_json=columns_json,
        s3_key=storage_key if storage_type == "s3" else None,
        storage_type=storage_type,
        file_content=None,  # streamed files never go through Postgres
        organisation_id=current_user.organisation_id,
        uploaded_by=current_user.id,
    )
    # Linked-sync bookkeeping
    if body.keep_linked:
        db_file.sharepoint_drive_id = body.drive_id
        db_file.sharepoint_item_id = body.item_id
        db_file.source_url = f"sharepoint://{body.drive_id}/{body.item_id}"
        db_file.last_synced_at = datetime.utcnow()
        db_file.sync_frequency = sf
    db.add(db_file)
    db.commit()

    # Register the auto-sync job if requested
    if body.keep_linked and sf != "off":
        try:
            import scheduler as app_scheduler
            app_scheduler.register_sharepoint_sync_job(db_file)
        except Exception:
            pass

    return {
        "id":            file_id,
        "filename":      body.filename,
        "file_size":     total_bytes,
        "row_count":     row_count,
        "column_count":  column_count,
        "linked":        body.keep_linked,
        "sync_frequency": sf,
        "message":       f"'{body.filename}' imported ({total_bytes // 1024} KB).",
    }


# ── Linked file management ────────────────────────────────────────────────────

@router.get("/linked-files")
async def list_linked_files(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Every linked SP/OD file in this org — the ones that can be resynced."""
    rows = db.query(DataFile).filter(
        DataFile.organisation_id == current_user.organisation_id,
        DataFile.sharepoint_item_id.isnot(None),
    ).all()
    return {
        "items": [
            {
                "file_id":        r.id,
                "filename":       r.original_filename,
                "size":           r.file_size,
                "drive_id":       r.sharepoint_drive_id,
                "item_id":        r.sharepoint_item_id,
                "frequency":      getattr(r, "sync_frequency", "off") or "off",
                "last_synced_at": r.last_synced_at.isoformat() if r.last_synced_at else None,
                "last_sync_error": getattr(r, "last_sync_error", None),
            }
            for r in rows
        ],
    }


class SyncScheduleRequest(BaseModel):
    frequency: str = Field(..., description="off, hourly, or daily")


@router.post("/{file_id}/sync-schedule")
async def set_sync_schedule(
    file_id: str,
    body: SyncScheduleRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    freq = (body.frequency or "off").lower().strip()
    if freq not in VALID_SYNC_FREQUENCIES:
        raise HTTPException(
            status_code=400,
            detail=f"frequency must be one of: {', '.join(sorted(VALID_SYNC_FREQUENCIES))}",
        )

    db_file = db.query(DataFile).filter(
        DataFile.id == file_id,
        DataFile.organisation_id == current_user.organisation_id,
    ).first()
    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")
    if not db_file.sharepoint_item_id:
        raise HTTPException(
            status_code=400,
            detail="This file is not linked to SharePoint. Re-import with keep_linked=true to enable auto-sync.",
        )

    db_file.sync_frequency = freq
    db.commit()
    db.refresh(db_file)

    try:
        import scheduler as app_scheduler
        app_scheduler.register_sharepoint_sync_job(db_file)
    except Exception:
        pass

    return {
        "file_id": file_id,
        "frequency": freq,
        "message": "Auto-sync disabled." if freq == "off" else f"Auto-sync set to {freq}.",
    }


@router.post("/{file_id}/resync")
async def resync_file(
    file_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Re-stream the latest content from SharePoint for a linked file."""
    db_file = db.query(DataFile).filter(
        DataFile.id == file_id,
        DataFile.organisation_id == current_user.organisation_id,
    ).first()
    if not db_file:
        raise HTTPException(status_code=404, detail="File not found")
    if not (db_file.sharepoint_drive_id and db_file.sharepoint_item_id):
        raise HTTPException(
            status_code=400,
            detail="This file is not linked to SharePoint. Re-import with keep_linked=true to enable resync.",
        )

    ext = ("." + db_file.original_filename.rsplit(".", 1)[-1].lower()) if "." in db_file.original_filename else ""
    if ext not in ALLOWED_EXT:
        raise HTTPException(status_code=400, detail="Unsupported file extension for resync.")

    token = await _get_valid_access_token(db, current_user.organisation_id)
    spooled, total_bytes = await _stream_graph_file_to_temp(
        token, db_file.sharepoint_drive_id, db_file.sharepoint_item_id,
    )
    try:
        row_count, column_count, columns_json = _parse_headers_from_temp(spooled, ext)
        # Overwrite in place: new storage_key (so viewers don't hit half-written files)
        new_key, storage_type, _ = _persist_temp_file(spooled, db_file.original_filename)
    finally:
        spooled.close()

    # Best-effort clean-up of the old blob. We don't care if it fails — R2/local
    # janitors can scoop up orphans later.
    old_key = db_file.filename
    try:
        if db_file.storage_type == "s3":
            from routers.files import delete_file_r2
            delete_file_r2(old_key)
        elif db_file.storage_type == "local":
            path = os.path.join(settings.LOCAL_UPLOAD_DIR, old_key)
            if os.path.exists(path):
                os.remove(path)
    except Exception:
        pass

    db_file.filename = new_key
    db_file.s3_key = new_key if storage_type == "s3" else None
    db_file.storage_type = storage_type
    db_file.file_size = total_bytes
    db_file.row_count = row_count
    db_file.column_count = column_count
    db_file.columns_json = columns_json
    db_file.last_synced_at = datetime.utcnow()
    db_file.last_sync_error = None
    db.commit()

    return {
        "file_id": file_id,
        "filename": db_file.original_filename,
        "size": total_bytes,
        "rows": row_count,
        "columns": column_count,
        "last_synced_at": db_file.last_synced_at.isoformat(),
        "message": f"Resynced — {total_bytes // 1024} KB, {row_count:,} rows.",
    }


# ══════════════════════════════════════════════════════════════════════════════
# SCHEDULER CALLBACK — invoked by APScheduler
# ══════════════════════════════════════════════════════════════════════════════

def execute_sharepoint_resync(file_id: str) -> None:
    """Resync a linked SharePoint DataFile. Called by APScheduler.
    Catches its own errors and stores them on last_sync_error."""
    from database import SessionLocal
    import asyncio
    import logging
    logger = logging.getLogger(__name__)

    db = SessionLocal()
    try:
        db_file = db.query(DataFile).filter(DataFile.id == file_id).first()
        if not db_file or not db_file.sharepoint_item_id:
            return
        if getattr(db_file, "sync_frequency", "off") == "off":
            return

        ext = ("." + db_file.original_filename.rsplit(".", 1)[-1].lower()) if "." in db_file.original_filename else ""
        if ext not in ALLOWED_EXT:
            return

        try:
            async def _go():
                token = await _get_valid_access_token(db, db_file.organisation_id)
                return await _stream_graph_file_to_temp(
                    token, db_file.sharepoint_drive_id, db_file.sharepoint_item_id,
                )
            spooled, total_bytes = asyncio.run(_go())
            try:
                row_count, column_count, columns_json = _parse_headers_from_temp(spooled, ext)
                new_key, storage_type, _ = _persist_temp_file(spooled, db_file.original_filename)
            finally:
                spooled.close()

            old_key = db_file.filename
            try:
                if db_file.storage_type == "s3":
                    from routers.files import delete_file_r2
                    delete_file_r2(old_key)
                elif db_file.storage_type == "local":
                    path = os.path.join(settings.LOCAL_UPLOAD_DIR, old_key)
                    if os.path.exists(path):
                        os.remove(path)
            except Exception:
                pass

            db_file.filename = new_key
            db_file.s3_key = new_key if storage_type == "s3" else None
            db_file.storage_type = storage_type
            db_file.file_size = total_bytes
            db_file.row_count = row_count
            db_file.column_count = column_count
            db_file.columns_json = columns_json
            db_file.last_synced_at = datetime.utcnow()
            db_file.last_sync_error = None
            db.commit()
            logger.info("SharePoint resync OK for %s (%s bytes, %s rows)", file_id, total_bytes, row_count)
        except Exception as exc:
            db.rollback()
            db_file.last_sync_error = str(exc)[:500]
            db.commit()
            logger.warning("SharePoint resync failed for %s: %s", file_id, exc)
    finally:
        db.close()
