from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from database import get_db, User, DataFile
from auth_utils import get_current_user
from config import settings
from pydantic import BaseModel
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, date, timedelta
import os
import json
import base64
import math
import statistics

router = APIRouter()

def load_file_data(file_id: str, org_id: str, db: Session):
    f = db.query(DataFile).filter(DataFile.id == file_id, DataFile.organisation_id == org_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")

    # Storage mode resolution — three modes supported:
    #   1. Postgres BYTEA (file_content) — small, legacy / dev mode
    #   2. Local disk (LOCAL_UPLOAD_DIR / filename) — when R2 isn't configured
    #   3. R2/S3 (storage_type='s3', s3_key set) — large files + streamed
    #      SharePoint imports
    # Until 2026-04-26 this function silently 404'd on mode 3 with "File
    # content not available. Please re-upload." That broke AI Insights,
    # Auto Report, Ask Your Data, RFM, Forecasting, KPIs, Summary, Preview,
    # and the new tool-use loop — basically every analytics path — for any
    # org using R2/SharePoint linked imports. Now we route through the
    # existing get_file_r2 helper.
    raw = None
    if f.file_content:
        raw = bytes(f.file_content)
    elif getattr(f, "storage_type", None) == "s3" and getattr(f, "s3_key", None):
        try:
            from routers.files import get_file_r2
            raw = get_file_r2(f.s3_key)
        except Exception as exc:
            # Surface the boto3 error message rather than the misleading
            # "please re-upload" — operators need to see auth/network
            # issues in the response, not have users blame their files.
            raise HTTPException(
                status_code=500,
                detail=f"Could not fetch file from object storage: {exc}",
            )
    else:
        path = os.path.join(settings.LOCAL_UPLOAD_DIR, f.filename)
        if os.path.exists(path):
            with open(path, "rb") as fp:
                raw = fp.read()
    if not raw:
        raise HTTPException(status_code=404, detail="File content not available. Please re-upload the file.")
    ext = os.path.splitext(f.original_filename)[1].lower()
    rows = []
    import io
    if ext in [".xlsx", ".xls"]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    elif ext == ".csv":
        import csv
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8", errors="ignore")))
        rows = list(reader)

    return rows, f.original_filename


@router.get("/preview/{file_id}")
def preview_file(file_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows, filename = load_file_data(file_id, current_user.organisation_id, db)
    if not rows:
        return {"filename": filename, "headers": [], "rows": []}
    headers = list(rows[0].keys())
    return {"filename": filename, "headers": headers, "rows": rows}

@router.post("/summary/{file_id}")
def get_summary(file_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows, filename = load_file_data(file_id, current_user.organisation_id, db)

    if not rows:
        return {"filename": filename, "rows": 0, "columns": [], "summary": {}}

    headers = list(rows[0].keys())
    summary = {}

    for h in headers:
        vals = [r[h] for r in rows if r[h] is not None and r[h] != ""]
        try:
            nums = [float(v) for v in vals]
            summary[h] = {
                "type": "numeric",
                "count": len(nums),
                "sum": round(sum(nums), 2),
                "mean": round(sum(nums) / len(nums), 2) if nums else 0,
                "min": round(min(nums), 2) if nums else 0,
                "max": round(max(nums), 2) if nums else 0
            }
        except (ValueError, TypeError):
            unique = list(set(str(v) for v in vals))
            summary[h] = {
                "type": "text",
                "count": len(vals),
                "unique": len(unique),
                "top_values": unique[:10]
            }

    return {"filename": filename, "rows": len(rows), "columns": len(headers), "summary": summary}

@router.post("/kpi/{file_id}")
def get_kpis(file_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    rows, filename = load_file_data(file_id, current_user.organisation_id, db)

    if not rows:
        return []

    headers = list(rows[0].keys())
    kpis = []

    for h in headers:
        try:
            nums = [float(r[h]) for r in rows if r[h] is not None and r[h] != ""]
            if nums:
                kpis.append({
                    "column": h,
                    "sum": round(sum(nums), 2),
                    "mean": round(sum(nums) / len(nums), 2),
                    "min": round(min(nums), 2),
                    "max": round(max(nums), 2),
                    "count": len(nums)
                })
        except (ValueError, TypeError):
            pass

    return kpis


# -----------------------------------------------------------------------------
# Forecasting 2.0 — pure-Python Holt-Winters + auto-method selection + CI bands
# -----------------------------------------------------------------------------
# Supports four models, picks the best via rolling backtest MAPE, and returns
# 80% / 95% prediction intervals using residual standard error.

class ForecastRequest(BaseModel):
    value_column: str
    date_column: Optional[str] = None
    periods: int = 6
    # "auto" | "ses" | "holt" | "holt_winters" | "linear"
    method: Optional[str] = "auto"
    # "auto" | "daily" | "weekly" | "monthly" | "quarterly" — only used with date_column
    aggregation: Optional[str] = "auto"


def _try_parse_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    fmts = (
        "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y",
        "%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S", "%b %d %Y", "%d %b %Y", "%Y",
    )
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except Exception:
        return None


def _period_key(d: date, agg: str) -> Tuple[int, int, int]:
    """Return a sortable key for the bucket containing d."""
    if agg == "daily":
        return (d.year, d.month, d.day)
    if agg == "weekly":
        iso = d.isocalendar()
        return (iso[0], iso[1], 0)
    if agg == "quarterly":
        q = (d.month - 1) // 3 + 1
        return (d.year, q, 0)
    # monthly
    return (d.year, d.month, 0)


def _label_for_key(key: Tuple[int, int, int], agg: str) -> str:
    y, a, b = key
    if agg == "daily":
        return f"{y:04d}-{a:02d}-{b:02d}"
    if agg == "weekly":
        return f"{y}-W{a:02d}"
    if agg == "quarterly":
        return f"{y} Q{a}"
    return f"{y}-{a:02d}"


def _next_key(key: Tuple[int, int, int], agg: str) -> Tuple[int, int, int]:
    y, a, b = key
    if agg == "daily":
        d = date(y, a, b) + timedelta(days=1)
        return (d.year, d.month, d.day)
    if agg == "weekly":
        if a >= 52:
            return (y + 1, 1, 0)
        return (y, a + 1, 0)
    if agg == "quarterly":
        if a >= 4:
            return (y + 1, 1, 0)
        return (y, a + 1, 0)
    if a >= 12:
        return (y + 1, 1, 0)
    return (y, a + 1, 0)


def _pick_aggregation(dates: List[date]) -> str:
    """Infer a sensible aggregation from the span + count of distinct dates."""
    if not dates:
        return "monthly"
    span_days = (max(dates) - min(dates)).days + 1
    n = len(set(dates))
    if span_days <= 60 and n > 20:
        return "daily"
    if span_days <= 400 and span_days > 60:
        return "weekly" if n > 40 else "monthly"
    # Long spans: prefer monthly when there's enough distinct data points
    # to resolve monthly buckets (24+), else fall back to quarterly.
    if span_days > 2200 and n < 30:
        return "quarterly"
    return "monthly"


def _aggregate_series(
    rows: List[dict], value_col: str, date_col: Optional[str], agg: str
) -> Tuple[List[Tuple[Tuple[int, int, int], float]], str]:
    """Return sorted list of (period_key, value) and the aggregation used."""
    if date_col:
        paired: List[Tuple[date, float]] = []
        for r in rows:
            d = _try_parse_date(r.get(date_col))
            if d is None:
                continue
            try:
                v = float(r.get(value_col))
            except (TypeError, ValueError):
                continue
            paired.append((d, v))
        if not paired:
            return [], agg
        if agg == "auto":
            agg = _pick_aggregation([d for d, _ in paired])
        buckets: Dict[Tuple[int, int, int], float] = {}
        for d, v in paired:
            k = _period_key(d, agg)
            buckets[k] = buckets.get(k, 0.0) + v
        series = sorted(buckets.items())
        return series, agg
    else:
        # No date column — use row order, label as period index.
        series = []
        for i, r in enumerate(rows):
            try:
                v = float(r.get(value_col))
                series.append(((i + 1, 0, 0), v))
            except (TypeError, ValueError):
                continue
        return series, "index"


# ---- Forecasting models ------------------------------------------------------

def _forecast_ses(y: List[float], h: int, alpha: float = 0.3) -> Tuple[List[float], List[float]]:
    """Simple exponential smoothing. Returns (forecast, in_sample_residuals)."""
    if not y:
        return [], []
    level = y[0]
    residuals = []
    for v in y[1:]:
        residuals.append(v - level)
        level = alpha * v + (1 - alpha) * level
    forecast = [level] * h
    return forecast, residuals


def _forecast_holt(y: List[float], h: int, alpha: float = 0.3, beta: float = 0.1) -> Tuple[List[float], List[float]]:
    """Holt's linear trend method."""
    n = len(y)
    if n < 2:
        return [y[0] if y else 0.0] * h, []
    level = y[0]
    trend = y[1] - y[0]
    residuals = []
    for v in y[1:]:
        fitted = level + trend
        residuals.append(v - fitted)
        new_level = alpha * v + (1 - alpha) * (level + trend)
        trend = beta * (new_level - level) + (1 - beta) * trend
        level = new_level
    forecast = [level + (i + 1) * trend for i in range(h)]
    return forecast, residuals


def _forecast_holt_winters(
    y: List[float], h: int, m: int,
    alpha: float = 0.3, beta: float = 0.1, gamma: float = 0.2,
) -> Tuple[List[float], List[float]]:
    """Holt-Winters additive seasonality. Requires y with len >= 2*m."""
    n = len(y)
    if n < 2 * m or m < 2:
        return _forecast_holt(y, h, alpha, beta)
    # Initialise level = mean of first season; trend = avg slope between first 2 seasons
    first_season = y[:m]
    second_season = y[m:2 * m]
    level = sum(first_season) / m
    trend = (sum(second_season) / m - sum(first_season) / m) / m
    seasonals = [first_season[i] - level for i in range(m)]
    residuals = []
    for i in range(n):
        v = y[i]
        s_idx = i % m
        fitted = level + trend + seasonals[s_idx]
        if i >= m:  # skip initialisation period in residual pool
            residuals.append(v - fitted)
        new_level = alpha * (v - seasonals[s_idx]) + (1 - alpha) * (level + trend)
        trend = beta * (new_level - level) + (1 - beta) * trend
        seasonals[s_idx] = gamma * (v - new_level) + (1 - gamma) * seasonals[s_idx]
        level = new_level
    forecast = [level + (i + 1) * trend + seasonals[(n + i) % m] for i in range(h)]
    return forecast, residuals


def _forecast_linear(y: List[float], h: int) -> Tuple[List[float], List[float]]:
    n = len(y)
    if n < 2:
        return [y[0] if y else 0.0] * h, []
    mx = (n - 1) / 2
    my = sum(y) / n
    num = sum((i - mx) * (v - my) for i, v in enumerate(y))
    den = sum((i - mx) ** 2 for i in range(n)) or 1e-9
    slope = num / den
    intercept = my - slope * mx
    residuals = [y[i] - (slope * i + intercept) for i in range(n)]
    forecast = [slope * (n + i) + intercept for i in range(h)]
    return forecast, residuals


def _detect_season_length(agg: str, n: int) -> int:
    """Best-guess seasonal period for the aggregation."""
    if agg == "daily":
        return 7 if n >= 14 else 0
    if agg == "weekly":
        return 52 if n >= 104 else (4 if n >= 8 else 0)
    if agg == "monthly":
        return 12 if n >= 24 else 0
    if agg == "quarterly":
        return 4 if n >= 8 else 0
    return 0


def _mape(actual: List[float], predicted: List[float]) -> float:
    pairs = [(a, p) for a, p in zip(actual, predicted) if abs(a) > 1e-9]
    if not pairs:
        return float("inf")
    return sum(abs(a - p) / abs(a) for a, p in pairs) * 100 / len(pairs)


def _rmse(actual: List[float], predicted: List[float]) -> float:
    if not actual:
        return 0.0
    return math.sqrt(sum((a - p) ** 2 for a, p in zip(actual, predicted)) / len(actual))


def _backtest(y: List[float], method: str, m: int, hold_out: int) -> Dict[str, float]:
    if len(y) <= hold_out + 3:
        return {"mape": float("inf"), "rmse": float("inf")}
    train = y[:-hold_out]
    test = y[-hold_out:]
    if method == "ses":
        pred, _ = _forecast_ses(train, hold_out)
    elif method == "holt":
        pred, _ = _forecast_holt(train, hold_out)
    elif method == "holt_winters":
        pred, _ = _forecast_holt_winters(train, hold_out, m)
    else:
        pred, _ = _forecast_linear(train, hold_out)
    return {"mape": _mape(test, pred), "rmse": _rmse(test, pred)}


@router.post("/forecast/{file_id}")
def forecast(
    file_id: str,
    body: ForecastRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows, filename = load_file_data(file_id, current_user.organisation_id, db)
    if not rows:
        raise HTTPException(status_code=400, detail="File has no rows")
    if body.value_column not in rows[0]:
        raise HTTPException(status_code=400, detail=f"Column '{body.value_column}' not found")
    if body.date_column and body.date_column not in rows[0]:
        raise HTTPException(status_code=400, detail=f"Date column '{body.date_column}' not found")

    h = max(1, min(48, body.periods or 6))

    series, agg_used = _aggregate_series(rows, body.value_column, body.date_column, body.aggregation or "auto")
    if len(series) < 4:
        raise HTTPException(status_code=400, detail="Need at least 4 data points to forecast")

    keys = [k for k, _ in series]
    y = [v for _, v in series]
    n = len(y)

    m = _detect_season_length(agg_used, n)

    # Auto-select method via rolling backtest
    hold_out = max(3, min(n // 4, 2 * m if m else 6))
    candidates = ["linear", "ses", "holt"]
    if m >= 2 and n >= 2 * m + hold_out:
        candidates.append("holt_winters")

    backtests = {c: _backtest(y, c, m, hold_out) for c in candidates}

    requested = (body.method or "auto").lower()
    if requested in candidates:
        chosen = requested
    elif requested == "auto":
        chosen = min(candidates, key=lambda c: backtests[c]["mape"])
    else:
        chosen = "linear"

    # Final fit + forecast on full series
    if chosen == "ses":
        forecast_vals, residuals = _forecast_ses(y, h)
    elif chosen == "holt":
        forecast_vals, residuals = _forecast_holt(y, h)
    elif chosen == "holt_winters":
        forecast_vals, residuals = _forecast_holt_winters(y, h, m)
    else:
        forecast_vals, residuals = _forecast_linear(y, h)

    # Residual std -> prediction intervals, widening with sqrt(step)
    sigma = statistics.stdev(residuals) if len(residuals) >= 2 else (statistics.pstdev(y) if len(y) >= 2 else 0.0)
    z80 = 1.2816
    z95 = 1.96
    lower_80, upper_80, lower_95, upper_95 = [], [], [], []
    for i, fv in enumerate(forecast_vals, start=1):
        band = sigma * math.sqrt(i)
        lower_80.append(fv - z80 * band)
        upper_80.append(fv + z80 * band)
        lower_95.append(fv - z95 * band)
        upper_95.append(fv + z95 * band)

    # Build response series
    historical = [
        {"label": _label_for_key(k, agg_used) if agg_used != "index" else f"#{k[0]}", "value": round(v, 4)}
        for k, v in series
    ]
    # Future labels
    future_labels = []
    k = keys[-1]
    for _ in range(h):
        k = _next_key(k, agg_used) if agg_used != "index" else (k[0] + 1, 0, 0)
        future_labels.append(_label_for_key(k, agg_used) if agg_used != "index" else f"#{k[0]}")

    forecast_out = [
        {
            "label": future_labels[i],
            "forecast": round(forecast_vals[i], 4),
            "lower_80": round(lower_80[i], 4),
            "upper_80": round(upper_80[i], 4),
            "lower_95": round(lower_95[i], 4),
            "upper_95": round(upper_95[i], 4),
        }
        for i in range(h)
    ]

    # Trend direction + summary stat
    first_half_mean = sum(y[: n // 2]) / max(1, n // 2)
    second_half_mean = sum(y[n // 2:]) / max(1, n - n // 2)
    pct_change = ((second_half_mean - first_half_mean) / first_half_mean * 100) if first_half_mean else 0.0

    method_labels = {
        "linear": "Linear Trend",
        "ses": "Simple Exponential Smoothing",
        "holt": "Holt's Linear Trend",
        "holt_winters": "Holt-Winters (Seasonal)",
    }

    return {
        "filename": filename,
        "value_column": body.value_column,
        "date_column": body.date_column,
        "aggregation": agg_used,
        "periods": h,
        "method": chosen,
        "method_label": method_labels.get(chosen, chosen),
        "season_length": m if chosen == "holt_winters" else 0,
        "n_points": n,
        "hold_out": hold_out,
        "backtest": {c: {"mape": round(backtests[c]["mape"], 2), "rmse": round(backtests[c]["rmse"], 4)} for c in candidates},
        "historical": historical,
        "forecast": forecast_out,
        "summary": {
            "last_actual": round(y[-1], 4),
            "next_forecast": round(forecast_vals[0], 4),
            "mape": round(backtests[chosen]["mape"], 2),
            "rmse": round(backtests[chosen]["rmse"], 4),
            "trend_direction": "up" if pct_change > 2 else ("down" if pct_change < -2 else "flat"),
            "trend_pct": round(pct_change, 2),
            "sigma": round(sigma, 4),
        },
    }
