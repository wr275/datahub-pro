"""Scheduled Reports 2.0 — CRUD + APScheduler job executor.

2.0 improvements on top of the 1.0 router:

* **Report templates** — `data_summary` (the old default), `kpi_digest`
  (numeric KPIs only), `alerts_digest` (variance / anomaly callouts),
  `raw_attachment` (minimal body, attach the source file as CSV).
* **Delivery log** — every send attempt (scheduled + manual + retries) is
  persisted as a `ScheduledReportDelivery` row. A GET /{id}/deliveries
  endpoint exposes the history to the UI.
* **Retry with backoff** — transient SendGrid failures get re-queued in a
  background thread up to `max_retries` times, with exponential backoff.
  Each attempt writes its own delivery-log row so the UI can trace the
  full chain.
* **CSV attachments** — optional. When `attach_csv=True` and a file is
  linked, the source data is attached as `<filename>.csv`.

Routes (mounted at /api/scheduled-reports):
    GET    /                    list all for org (+ next_run_at + recent delivery summary)
    POST   /                    create
    PUT    /{id}                update
    DELETE /{id}                delete (cascades delivery log)
    PATCH  /{id}/toggle         pause / resume
    POST   /{id}/send-now       immediate manual trigger (returns 202)
    GET    /{id}/deliveries     paged delivery log (newest first)
    GET    /templates           returns the available template catalogue
"""
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import base64
import uuid
import io
import os
import csv
import statistics
import logging
import threading
import time

from database import (
    get_db,
    ScheduledReport,
    ScheduledReportDelivery,
    DataFile,
    User,
)
from auth_utils import get_current_user
from config import settings

logger = logging.getLogger(__name__)
router = APIRouter()


# ─── Template catalogue ──────────────────────────────────────────────────────

TEMPLATES = [
    {
        "id": "data_summary",
        "name": "Data summary",
        "desc": "Numeric stats for up to 5 columns plus file metadata. Good default.",
        "needs_file": True,
    },
    {
        "id": "kpi_digest",
        "name": "KPI digest",
        "desc": "Single-number headlines for each numeric column — compact one-page email.",
        "needs_file": True,
    },
    {
        "id": "alerts_digest",
        "name": "Alerts digest",
        "desc": "Flags columns with wide swings or outliers (mean ±2σ). Nothing to report = quieter email.",
        "needs_file": True,
    },
    {
        "id": "raw_attachment",
        "name": "Raw file (attachment only)",
        "desc": "Minimal body, attach the linked file as CSV. Best for sending the dataset to analysts.",
        "needs_file": True,
    },
]

TEMPLATE_IDS = {t["id"] for t in TEMPLATES}


# ─── Schemas ─────────────────────────────────────────────────────────────────

class ScheduledReportCreate(BaseModel):
    name: str
    report_type: str = "data_summary"
    frequency: str               # daily | weekly | monthly
    day_of_week: Optional[str] = None
    day_of_month: Optional[int] = None
    send_time: str = "08:00"
    recipients: str              # comma-separated
    file_id: Optional[str] = None
    attach_csv: bool = False
    max_retries: int = 2


class ScheduledReportUpdate(BaseModel):
    name: Optional[str] = None
    report_type: Optional[str] = None
    frequency: Optional[str] = None
    day_of_week: Optional[str] = None
    day_of_month: Optional[int] = None
    send_time: Optional[str] = None
    recipients: Optional[str] = None
    file_id: Optional[str] = None
    attach_csv: Optional[bool] = None
    max_retries: Optional[int] = None


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _next_run_at(schedule: ScheduledReport) -> Optional[str]:
    """Ask APScheduler what this job's next fire time is. Returns ISO or None."""
    try:
        from scheduler import _scheduler  # noqa: F401 — internal helper
        import scheduler as s
        if s._scheduler is None:
            return None
        job = s._scheduler.get_job(schedule.id)
        if not job or job.next_run_time is None:
            return None
        return job.next_run_time.astimezone(timezone.utc).isoformat()
    except Exception:
        return None


def _delivery_summary(db: Session, report_id: str) -> dict:
    """Return the most-recent delivery + counts for list/detail responses."""
    latest = (
        db.query(ScheduledReportDelivery)
        .filter(ScheduledReportDelivery.report_id == report_id)
        .order_by(ScheduledReportDelivery.attempted_at.desc())
        .first()
    )
    total = db.query(ScheduledReportDelivery).filter(ScheduledReportDelivery.report_id == report_id).count()
    failed = (
        db.query(ScheduledReportDelivery)
        .filter(
            ScheduledReportDelivery.report_id == report_id,
            ScheduledReportDelivery.status == "failed",
        )
        .count()
    )
    return {
        "last_status": latest.status if latest else None,
        "last_error": latest.error_message if latest else None,
        "last_attempted_at": latest.attempted_at.isoformat() if latest and latest.attempted_at else None,
        "total_attempts": total,
        "failed_attempts": failed,
    }


def _serialize(s: ScheduledReport, db: Optional[Session] = None) -> dict:
    base = {
        "id": s.id,
        "name": s.name,
        "report_type": s.report_type,
        "frequency": s.frequency,
        "day_of_week": s.day_of_week,
        "day_of_month": s.day_of_month,
        "send_time": s.send_time,
        "recipients": s.recipients,
        "file_id": s.file_id,
        "attach_csv": bool(getattr(s, "attach_csv", False)),
        "max_retries": int(getattr(s, "max_retries", 2) or 0),
        "status": s.status,
        "last_run_at": s.last_run_at.isoformat() if s.last_run_at else None,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "next_run_at": _next_run_at(s),
    }
    if db is not None:
        base["deliveries"] = _delivery_summary(db, s.id)
    return base


def _parse_file(f: DataFile) -> tuple[list, list, bytes]:
    """Return (headers, rows, raw_bytes) from a DataFile. Rows capped at 500
    for stats; raw_bytes is the full file (for attachments)."""
    from routers.utils import load_file_bytes
    raw = load_file_bytes(f) or b""
    ext = os.path.splitext(f.original_filename or "")[1].lower()

    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return [], [], raw
        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(header_row)]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=501, values_only=True):
            rows.append({headers[j]: (v if v is not None else "") for j, v in enumerate(row)})
        wb.close()
        return headers, rows, raw

    text = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows = [dict(r) for i, r in enumerate(reader) if i < 500]
    return headers, rows, raw


def _numeric_column(rows: list, col: str) -> list:
    out = []
    for row in rows:
        try:
            out.append(float(str(row.get(col, "")).replace(",", "").replace("$", "").strip()))
        except (ValueError, TypeError):
            pass
    return out


def _compute_stats(headers: list, rows: list, limit: int = 5) -> dict:
    """Basic numeric stats for the first `limit` numeric columns."""
    stats = {}
    if not rows:
        return stats
    for col in headers:
        if len(stats) >= limit:
            break
        values = _numeric_column(rows, col)
        if len(values) < 3:
            continue
        stats[col] = {
            "min":   f"{min(values):,.2f}",
            "max":   f"{max(values):,.2f}",
            "mean":  f"{statistics.mean(values):,.2f}",
            "count": len(values),
        }
    return stats


def _compute_alerts(headers: list, rows: list) -> list:
    """Scan numeric columns for outliers (values > mean + 2σ).

    Returns a list of {column, count, threshold, top_value} — empty when nothing
    notable. Helps power the `alerts_digest` template without pulling in
    heavier anomaly ML.
    """
    alerts = []
    for col in headers:
        vals = _numeric_column(rows, col)
        if len(vals) < 10:
            continue
        try:
            mean = statistics.mean(vals)
            sd = statistics.pstdev(vals)
        except statistics.StatisticsError:
            continue
        if sd == 0:
            continue
        threshold = mean + 2 * sd
        outliers = [v for v in vals if v > threshold]
        if outliers:
            alerts.append({
                "column": col,
                "count": len(outliers),
                "threshold": f"{threshold:,.2f}",
                "top_value": f"{max(outliers):,.2f}",
                "mean": f"{mean:,.2f}",
            })
    # Sort by highest outlier count first.
    alerts.sort(key=lambda a: -a["count"])
    return alerts[:6]


# ─── Email body builders ─────────────────────────────────────────────────────

def _header_block(schedule: ScheduledReport) -> str:
    now_str = datetime.now(timezone.utc).strftime("%A, %d %B %Y at %H:%M UTC")
    freq_label = (schedule.frequency or "").capitalize()
    if schedule.frequency == "weekly":
        freq_label = f"Weekly · {schedule.day_of_week or 'Monday'}s"
    elif schedule.frequency == "monthly":
        freq_label = f"Monthly · Day {schedule.day_of_month or 1}"
    return f"""
    <div style="background:linear-gradient(135deg,#0c1446,#1a2a6c);padding:28px 32px;">
      <div style="margin-bottom:16px;">
        <span style="display:inline-block;width:32px;height:32px;background:#e91e8c;border-radius:7px;color:#fff;font-weight:800;text-align:center;line-height:32px;margin-right:8px;vertical-align:middle;">D</span>
        <span style="color:#fff;font-weight:700;font-size:1rem;opacity:0.9;vertical-align:middle;">DataHub Pro</span>
      </div>
      <div style="color:#fff;font-size:1.4rem;font-weight:800;margin-bottom:4px;">{schedule.name}</div>
      <div style="color:rgba(255,255,255,0.65);font-size:0.85rem;">{freq_label} &nbsp;·&nbsp; {now_str}</div>
    </div>
    """


def _footer_block() -> str:
    frontend_url = settings.FRONTEND_URL.split(",")[0].strip()
    return f"""
    <div style="text-align:center;margin-top:28px;">
      <a href="{frontend_url}/hub" style="display:inline-block;background:#e91e8c;color:#fff;font-weight:700;font-size:0.9rem;padding:13px 28px;border-radius:8px;text-decoration:none;">
        Open DataHub Pro →
      </a>
    </div>
    <div style="padding:18px 32px;background:#f9fafb;border-top:1px solid #e5e7eb;margin-top:28px;">
      <div style="color:#9ca3af;font-size:0.78rem;text-align:center;">
        You're receiving this because a scheduled report was set up for your account.<br>
        Manage schedules at <a href="{frontend_url}/scheduled-reports" style="color:#6b7280;">{frontend_url}/scheduled-reports</a>
      </div>
    </div>
    """


def _file_block(data_file: Optional[DataFile]) -> str:
    if not data_file:
        return """<div style="background:#fef3c7;border-radius:8px;padding:14px 18px;margin-bottom:24px;color:#92400e;font-size:0.875rem;">
          This schedule isn't linked to a data file. Add one in <em>Scheduled reports</em> to get automatic stats.
        </div>"""
    synced = data_file.last_synced_at or data_file.created_at
    synced_str = synced.strftime("%d %b %Y %H:%M UTC") if synced else "—"
    return f"""
    <div style="background:#f0f4ff;border-radius:8px;padding:16px 20px;margin-bottom:24px;">
      <div style="font-size:0.8rem;color:#6b7280;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.05em;">Data source</div>
      <div style="font-weight:700;color:#0c1446;font-size:1rem;">{data_file.original_filename}</div>
      <div style="color:#6b7280;font-size:0.85rem;margin-top:4px;">
        {data_file.row_count or 0:,} rows &nbsp;·&nbsp; {data_file.column_count or 0} columns
        &nbsp;·&nbsp; Last synced {synced_str}
      </div>
    </div>"""


def _stats_table(stats: dict) -> str:
    if not stats:
        return """<div style="background:#f9fafb;border-radius:8px;padding:16px 20px;color:#6b7280;font-size:0.875rem;margin-bottom:24px;">
          No numeric data found in the linked file.
        </div>"""
    body = ""
    for col, s in stats.items():
        body += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#374151;font-weight:500;">{col}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#6b7280;text-align:right;">{s['min']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#0c1446;font-weight:700;text-align:right;">{s['mean']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#6b7280;text-align:right;">{s['max']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#9ca3af;text-align:right;">{s['count']:,}</td>
        </tr>"""
    return f"""
    <div style="margin-bottom:24px;">
      <div style="font-size:0.8rem;color:#6b7280;margin-bottom:10px;text-transform:uppercase;letter-spacing:0.05em;">Key metrics</div>
      <table style="width:100%;border-collapse:collapse;font-size:0.875rem;">
        <thead>
          <tr style="background:#f9fafb;">
            <th style="padding:8px 14px;text-align:left;color:#374151;font-weight:600;border-bottom:2px solid #e5e7eb;">Column</th>
            <th style="padding:8px 14px;text-align:right;color:#374151;font-weight:600;border-bottom:2px solid #e5e7eb;">Min</th>
            <th style="padding:8px 14px;text-align:right;color:#374151;font-weight:600;border-bottom:2px solid #e5e7eb;">Average</th>
            <th style="padding:8px 14px;text-align:right;color:#374151;font-weight:600;border-bottom:2px solid #e5e7eb;">Max</th>
            <th style="padding:8px 14px;text-align:right;color:#374151;font-weight:600;border-bottom:2px solid #e5e7eb;">Rows</th>
          </tr>
        </thead>
        <tbody>{body}</tbody>
      </table>
    </div>"""


def _kpi_tiles(stats: dict) -> str:
    if not stats:
        return _stats_table(stats)
    tiles = ""
    for col, s in list(stats.items())[:6]:
        tiles += f"""
          <div style="flex:1 1 180px;background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;">
            <div style="font-size:0.72rem;color:#6b7280;text-transform:uppercase;letter-spacing:0.5px;font-weight:700;">{col}</div>
            <div style="font-size:1.4rem;font-weight:800;color:#0c1446;margin-top:2px;">{s['mean']}</div>
            <div style="font-size:0.72rem;color:#9ca3af;margin-top:4px;">avg over {s['count']:,} rows · min {s['min']} / max {s['max']}</div>
          </div>"""
    return f"""<div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px;">{tiles}</div>"""


def _alerts_section(alerts: list) -> str:
    if not alerts:
        return """
        <div style="background:#ecfdf5;border:1px solid #a7f3d0;border-radius:8px;padding:14px 18px;color:#065f46;font-size:0.875rem;margin-bottom:24px;">
          ✓ No outliers above mean + 2σ found in this cut. Nothing to action.
        </div>"""
    rows = ""
    for a in alerts:
        rows += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#374151;font-weight:500;">{a['column']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#ef4444;font-weight:700;text-align:right;">{a['count']:,}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#6b7280;text-align:right;">{a['threshold']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #f3f4f6;color:#0c1446;font-weight:700;text-align:right;">{a['top_value']}</td>
        </tr>"""
    return f"""
    <div style="margin-bottom:24px;">
      <div style="font-size:0.8rem;color:#6b7280;margin-bottom:10px;text-transform:uppercase;letter-spacing:0.05em;">Outliers flagged</div>
      <table style="width:100%;border-collapse:collapse;font-size:0.875rem;">
        <thead>
          <tr style="background:#fef2f2;">
            <th style="padding:8px 14px;text-align:left;color:#991b1b;font-weight:600;">Column</th>
            <th style="padding:8px 14px;text-align:right;color:#991b1b;font-weight:600;">Outliers</th>
            <th style="padding:8px 14px;text-align:right;color:#991b1b;font-weight:600;">Threshold</th>
            <th style="padding:8px 14px;text-align:right;color:#991b1b;font-weight:600;">Highest</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""


def _envelope(inner_body: str) -> str:
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
  <div style="max-width:640px;margin:32px auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
    {inner_body}
  </div>
</body></html>"""


def _build_html(schedule: ScheduledReport, data_file: Optional[DataFile],
                headers: list, rows: list) -> str:
    template = (schedule.report_type or "data_summary").lower()
    header = _header_block(schedule)
    file_block = _file_block(data_file)

    if template == "kpi_digest":
        stats = _compute_stats(headers, rows, limit=6)
        body = f"""<div style="padding:28px 32px;">{file_block}{_kpi_tiles(stats)}{_footer_block()}</div>"""
    elif template == "alerts_digest":
        alerts = _compute_alerts(headers, rows)
        body = f"""<div style="padding:28px 32px;">{file_block}{_alerts_section(alerts)}{_footer_block()}</div>"""
    elif template == "raw_attachment":
        # Minimal body — the payload is in the CSV attachment.
        rows_copy = f"{len(rows):,}+ rows" if rows else "0 rows"
        body = f"""<div style="padding:28px 32px;">{file_block}
          <div style="background:#f9fafb;border-radius:8px;padding:16px 20px;color:#374151;font-size:0.9rem;line-height:1.55;margin-bottom:24px;">
            Attached as CSV: <strong>{(data_file.original_filename if data_file else 'report')}</strong> ({rows_copy}).<br>
            Open in Excel, Numbers, or your tool of choice.
          </div>
          {_footer_block()}
        </div>"""
    else:  # data_summary (default)
        stats = _compute_stats(headers, rows, limit=5)
        body = f"""<div style="padding:28px 32px;">{file_block}{_stats_table(stats)}{_footer_block()}</div>"""

    return _envelope(header + body)


# ─── Email sender ────────────────────────────────────────────────────────────

def _send_email(to_list: list[str], subject: str, html: str,
                attachments: Optional[list[dict]] = None) -> tuple[bool, Optional[str]]:
    """Send via SendGrid. Returns (ok, error_message_or_None).

    `attachments` is a list of {filename, content (bytes), mime_type}.
    """
    if not settings.SENDGRID_API_KEY:
        return False, "SENDGRID_API_KEY not set"
    try:
        import sendgrid as sg_lib
        from sendgrid.helpers.mail import Mail, To, Attachment, FileContent, FileName, FileType, Disposition
        msg = Mail(
            from_email=settings.FROM_EMAIL,
            subject=subject,
            html_content=html,
        )
        msg.to = [To(email=e.strip()) for e in to_list if e.strip()]
        for att in attachments or []:
            encoded = base64.b64encode(att["content"]).decode()
            msg.attachment = Attachment(
                FileContent(encoded),
                FileName(att["filename"]),
                FileType(att.get("mime_type", "application/octet-stream")),
                Disposition("attachment"),
            )
        client = sg_lib.SendGridAPIClient(api_key=settings.SENDGRID_API_KEY)
        resp = client.send(msg)
        if resp.status_code in (200, 201, 202):
            return True, None
        return False, f"SendGrid returned HTTP {resp.status_code}"
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


# ─── Delivery log helpers ────────────────────────────────────────────────────

def _log_delivery(db: Session, *, report_id: str, status: str, attempt: int,
                  recipients: str, error: Optional[str], rows_included: Optional[int],
                  trigger: str) -> None:
    row = ScheduledReportDelivery(
        id=str(uuid.uuid4()),
        report_id=report_id,
        status=status,
        recipients=recipients,
        attempt=attempt,
        error_message=(error or None),
        rows_included=rows_included,
        trigger=trigger,
    )
    db.add(row)
    db.commit()


# ─── Job executor (called by APScheduler + Send-now) ─────────────────────────

def _execute_attempt(report_id: str, attempt: int, trigger: str) -> tuple[bool, Optional[str], int]:
    """Single send attempt. Returns (ok, error, rows_included)."""
    from database import SessionLocal
    db = SessionLocal()
    try:
        schedule = db.query(ScheduledReport).filter(ScheduledReport.id == report_id).first()
        if not schedule:
            return False, "schedule not found", 0
        if schedule.status != "active" and trigger == "scheduled":
            _log_delivery(db, report_id=report_id, status="skipped", attempt=attempt,
                          recipients=schedule.recipients, error="schedule is paused",
                          rows_included=0, trigger=trigger)
            return False, "paused", 0

        data_file = None
        headers: list = []
        rows: list = []
        raw_bytes: bytes = b""
        rows_count = 0

        if schedule.file_id:
            data_file = db.query(DataFile).filter(DataFile.id == schedule.file_id).first()
            if data_file:
                try:
                    headers, rows, raw_bytes = _parse_file(data_file)
                    rows_count = len(rows)
                except Exception as exc:
                    logger.warning("Could not parse file %s: %s", schedule.file_id, exc)
                    raw_bytes = b""

        html = _build_html(schedule, data_file, headers, rows)
        to_list = [e.strip() for e in (schedule.recipients or "").split(",") if e.strip()]
        if not to_list:
            err = "no recipients configured"
            _log_delivery(db, report_id=report_id, status="failed", attempt=attempt,
                          recipients="", error=err, rows_included=rows_count, trigger=trigger)
            return False, err, rows_count

        subject = f"{schedule.name} — {datetime.now(timezone.utc).strftime('%d %b %Y')}"

        attachments: list[dict] = []
        if bool(getattr(schedule, "attach_csv", False)) and data_file and raw_bytes:
            # Convert xlsx to csv if needed; CSV we send as-is.
            ext = os.path.splitext(data_file.original_filename or "")[1].lower()
            if ext in (".xlsx", ".xls"):
                buf = io.StringIO()
                writer = csv.writer(buf)
                if headers:
                    writer.writerow(headers)
                    for r in rows:
                        writer.writerow([r.get(h, "") for h in headers])
                csv_bytes = buf.getvalue().encode("utf-8")
                fname = (os.path.splitext(data_file.original_filename)[0] or "report") + ".csv"
                attachments.append({"filename": fname, "content": csv_bytes, "mime_type": "text/csv"})
            else:
                attachments.append({
                    "filename": data_file.original_filename or "report.csv",
                    "content": raw_bytes,
                    "mime_type": "text/csv",
                })

        ok, err = _send_email(to_list, subject, html, attachments)
        if ok:
            schedule.last_run_at = datetime.utcnow()
            db.commit()
            _log_delivery(db, report_id=report_id, status="sent", attempt=attempt,
                          recipients=schedule.recipients, error=None,
                          rows_included=rows_count, trigger=trigger)
            return True, None, rows_count

        _log_delivery(db, report_id=report_id, status="failed", attempt=attempt,
                      recipients=schedule.recipients, error=err,
                      rows_included=rows_count, trigger=trigger)
        return False, err, rows_count
    finally:
        db.close()


def execute_scheduled_report(report_id: str, *, trigger: str = "scheduled") -> None:
    """Top-level entrypoint — runs attempt(s) with exponential backoff on
    transient failures (up to `max_retries`). Called by:
        * APScheduler cron trigger (trigger='scheduled')
        * /send-now route (trigger='manual')
    """
    from database import SessionLocal
    db = SessionLocal()
    try:
        schedule = db.query(ScheduledReport).filter(ScheduledReport.id == report_id).first()
        max_retries = int(getattr(schedule, "max_retries", 2) or 0) if schedule else 2
    finally:
        db.close()

    attempt = 1
    while True:
        ok, err, _rows = _execute_attempt(report_id, attempt, trigger)
        if ok:
            logger.info("[sched %s] sent on attempt %d (trigger=%s)", report_id, attempt, trigger)
            return
        if attempt > max_retries:
            logger.error("[sched %s] failed after %d attempts: %s", report_id, attempt, err)
            return
        # Exponential backoff: 5s, 25s, 125s … capped.
        delay = min(300, 5 * (5 ** (attempt - 1)))
        logger.warning("[sched %s] attempt %d failed (%s) — retrying in %ds", report_id, attempt, err, delay)
        time.sleep(delay)
        attempt += 1


# ─── Routes ──────────────────────────────────────────────────────────────────

@router.get("/templates")
def list_templates(current_user: User = Depends(get_current_user)):
    """Expose the template catalogue to the frontend so new templates can be
    added server-side without a UI redeploy."""
    return {"templates": TEMPLATES}


@router.get("/")
def list_schedules(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(ScheduledReport)
        .filter(ScheduledReport.organisation_id == current_user.organisation_id)
        .order_by(ScheduledReport.created_at.desc())
        .all()
    )
    return [_serialize(r, db=db) for r in rows]


@router.post("/", status_code=201)
def create_schedule(
    req: ScheduledReportCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if req.report_type not in TEMPLATE_IDS:
        raise HTTPException(400, f"Unknown report_type '{req.report_type}'")
    if req.frequency not in ("daily", "weekly", "monthly"):
        raise HTTPException(400, "frequency must be daily, weekly, or monthly")
    if req.file_id:
        f = db.query(DataFile).filter(
            DataFile.id == req.file_id,
            DataFile.organisation_id == current_user.organisation_id,
        ).first()
        if not f:
            raise HTTPException(status_code=404, detail="File not found")
    if not (req.recipients or "").strip():
        raise HTTPException(400, "At least one recipient email is required")
    if req.max_retries < 0 or req.max_retries > 5:
        raise HTTPException(400, "max_retries must be between 0 and 5")

    schedule = ScheduledReport(
        id=str(uuid.uuid4()),
        name=req.name,
        report_type=req.report_type,
        frequency=req.frequency,
        day_of_week=req.day_of_week,
        day_of_month=req.day_of_month,
        send_time=req.send_time,
        recipients=req.recipients,
        file_id=req.file_id,
        attach_csv=bool(req.attach_csv),
        max_retries=int(req.max_retries),
        status="active",
        organisation_id=current_user.organisation_id,
        created_by=current_user.id,
    )
    db.add(schedule)
    db.commit()
    db.refresh(schedule)

    from scheduler import register_job
    register_job(schedule)

    return _serialize(schedule, db=db)


@router.put("/{schedule_id}")
def update_schedule(
    schedule_id: str,
    req: ScheduledReportUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    schedule = db.query(ScheduledReport).filter(
        ScheduledReport.id == schedule_id,
        ScheduledReport.organisation_id == current_user.organisation_id,
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    patch = req.model_dump(exclude_unset=True)
    if "report_type" in patch and patch["report_type"] not in TEMPLATE_IDS:
        raise HTTPException(400, f"Unknown report_type '{patch['report_type']}'")
    if "frequency" in patch and patch["frequency"] not in ("daily", "weekly", "monthly"):
        raise HTTPException(400, "frequency must be daily, weekly, or monthly")
    if "max_retries" in patch and (patch["max_retries"] < 0 or patch["max_retries"] > 5):
        raise HTTPException(400, "max_retries must be between 0 and 5")

    for field, value in patch.items():
        setattr(schedule, field, value)
    schedule.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(schedule)

    from scheduler import register_job, remove_job
    if schedule.status == "active":
        register_job(schedule)
    else:
        remove_job(schedule_id)

    return _serialize(schedule, db=db)


@router.delete("/{schedule_id}", status_code=204)
def delete_schedule(
    schedule_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    schedule = db.query(ScheduledReport).filter(
        ScheduledReport.id == schedule_id,
        ScheduledReport.organisation_id == current_user.organisation_id,
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    from scheduler import remove_job
    remove_job(schedule_id)

    # Explicitly purge delivery log rows — ON DELETE CASCADE should handle it,
    # but sqlite test envs occasionally don't enforce FKs.
    db.query(ScheduledReportDelivery).filter(ScheduledReportDelivery.report_id == schedule_id).delete()
    db.delete(schedule)
    db.commit()


@router.patch("/{schedule_id}/toggle")
def toggle_schedule(
    schedule_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    schedule = db.query(ScheduledReport).filter(
        ScheduledReport.id == schedule_id,
        ScheduledReport.organisation_id == current_user.organisation_id,
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    from scheduler import register_job, remove_job
    if schedule.status == "active":
        schedule.status = "paused"
        remove_job(schedule_id)
    else:
        schedule.status = "active"
        register_job(schedule)

    db.commit()
    db.refresh(schedule)
    return _serialize(schedule, db=db)


@router.post("/{schedule_id}/send-now", status_code=202)
def send_now(
    schedule_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    schedule = db.query(ScheduledReport).filter(
        ScheduledReport.id == schedule_id,
        ScheduledReport.organisation_id == current_user.organisation_id,
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    # Fire-and-forget background thread so the request returns immediately.
    # Retry/backoff is handled inside execute_scheduled_report().
    t = threading.Thread(
        target=execute_scheduled_report,
        args=(schedule_id,),
        kwargs={"trigger": "manual"},
        daemon=True,
    )
    t.start()
    return {"message": f"Report '{schedule.name}' queued for immediate delivery"}


@router.get("/{schedule_id}/deliveries")
def list_deliveries(
    schedule_id: str,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    schedule = db.query(ScheduledReport).filter(
        ScheduledReport.id == schedule_id,
        ScheduledReport.organisation_id == current_user.organisation_id,
    ).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    limit = max(1, min(200, int(limit or 50)))
    rows = (
        db.query(ScheduledReportDelivery)
        .filter(ScheduledReportDelivery.report_id == schedule_id)
        .order_by(ScheduledReportDelivery.attempted_at.desc())
        .limit(limit)
        .all()
    )
    return [{
        "id": r.id,
        "status": r.status,
        "attempt": r.attempt,
        "attempted_at": r.attempted_at.isoformat() if r.attempted_at else None,
        "recipients": r.recipients,
        "error_message": r.error_message,
        "rows_included": r.rows_included,
        "trigger": r.trigger,
    } for r in rows]
