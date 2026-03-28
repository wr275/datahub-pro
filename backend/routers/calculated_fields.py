from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db, User, DataFile, CalculatedFieldSet
from auth_utils import get_current_user
from pydantic import BaseModel
from typing import List, Optional
import json, uuid, io, csv, os

router = APIRouter()


# ── file loader (mirrors analytics.py) ─────────────────────────
def _load(file_id: str, org_id: str, db: Session):
    f = db.query(DataFile).filter(
        DataFile.id == file_id,
        DataFile.organisation_id == org_id,
    ).first()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")
    raw = None
    if f.file_content:
        raw = bytes(f.file_content)
    else:
        from config import settings
        path = os.path.join(settings.LOCAL_UPLOAD_DIR, f.filename)
        if os.path.exists(path):
            with open(path, "rb") as fp:
                raw = fp.read()
    if not raw:
        raise HTTPException(status_code=404, detail="File content not available. Please re-upload.")
    ext = os.path.splitext(f.original_filename)[1].lower()
    rows = []
    if ext in [".xlsx", ".xls"]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        import csv as csv_mod
        reader = csv_mod.DictReader(io.StringIO(raw.decode("utf-8", errors="ignore")))
        rows = list(reader)
    return rows, f.original_filename


# ── formula engine ──────────────────────────────────────────────
def _get_val(row, col, col_type, const):
    if col_type == "constant":
        return float(const) if const is not None else 0.0
    v = row.get(col)
    if v is None or v == "":
        return 0.0
    return float(v)


def _apply(rows, fields):
    result = []
    for row in rows:
        r = dict(row)
        for fld in fields:
            try:
                a = _get_val(r, fld.col_a, fld.col_a_type, fld.const_a)
                b = _get_val(r, fld.col_b, fld.col_b_type, fld.const_b)
                if fld.operator == "+":
                    v = a + b
                elif fld.operator == "-":
                    v = a - b
                elif fld.operator == "*":
                    v = a * b
                elif fld.operator == "/":
                    v = (a / b) if b != 0 else None
                elif fld.operator == "%":
                    v = ((a / b) * 100) if b != 0 else None
                else:
                    v = None
                r[fld.name] = round(v, 4) if v is not None else None
            except Exception:
                r[fld.name] = None
        result.append(r)
    return result


# ── Pydantic models ─────────────────────────────────────────────
class FieldDef(BaseModel):
    name: str
    col_a: str
    col_a_type: str = "column"   # "column" | "constant"
    const_a: Optional[float] = None
    operator: str                # +  -  *  /  %
    col_b: str
    col_b_type: str = "column"   # "column" | "constant"
    const_b: Optional[float] = None


class PreviewReq(BaseModel):
    file_id: str
    fields: List[FieldDef]


class SaveReq(BaseModel):
    set_name: str
    file_id: str
    fields: List[FieldDef]


# ── endpoints ───────────────────────────────────────────────────
@router.post("/preview")
def preview(
    req: PreviewReq,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows, filename = _load(req.file_id, user.organisation_id, db)
    enriched = _apply(rows, req.fields)
    headers = list(enriched[0].keys()) if enriched else []
    return {"filename": filename, "headers": headers, "rows": enriched[:50]}


@router.post("/export")
def export_csv(
    req: PreviewReq,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows, filename = _load(req.file_id, user.organisation_id, db)
    enriched = _apply(rows, req.fields)
    if not enriched:
        raise HTTPException(400, "No data to export")
    headers = list(enriched[0].keys())
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=headers)
    w.writeheader()
    w.writerows(enriched)
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="enriched_{filename}"'},
    )


@router.get("/")
def list_sets(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    sets = (
        db.query(CalculatedFieldSet)
        .filter(CalculatedFieldSet.organisation_id == user.organisation_id)
        .order_by(CalculatedFieldSet.created_at.desc())
        .all()
    )
    return [
        {
            "id": s.id,
            "name": s.name,
            "file_id": s.file_id,
            "fields": json.loads(s.fields_json),
            "created_at": str(s.created_at),
        }
        for s in sets
    ]


@router.post("/")
def save_set(
    req: SaveReq,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    s = CalculatedFieldSet(
        id=str(uuid.uuid4()),
        name=req.set_name,
        file_id=req.file_id,
        fields_json=json.dumps([f.dict() for f in req.fields]),
        organisation_id=user.organisation_id,
        created_by=user.id,
    )
    db.add(s)
    db.commit()
    return {"id": s.id, "message": "Field set saved"}


@router.delete("/{set_id}")
def delete_set(
    set_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    s = db.query(CalculatedFieldSet).filter(
        CalculatedFieldSet.id == set_id,
        CalculatedFieldSet.organisation_id == user.organisation_id,
    ).first()
    if not s:
        raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()
    return {"message": "Deleted"}
