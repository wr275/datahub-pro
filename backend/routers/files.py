from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
from database import get_db, User, DataFile, Organisation
from auth_utils import get_current_user
from config import settings
import uuid
import os
import json
import hashlib
from datetime import datetime

router = APIRouter()

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB supported via R2

# ── Sample data catalog ──────────────────────────────────────────────────────
# Bundled demo datasets shipped with the repo at backend/sample_data/. These
# power the "See a live dashboard with sample data" path in the zero-files
# onboarding flow — a new user can populate their workspace without first
# having to find and upload their own file.
#
# Each entry must map to a real file in SAMPLE_DATA_DIR. Column names are
# chosen to hit the ExecutiveDashboard auto-detection regexes (time column
# regex looks for month|date|period; primary metric regex looks for
# revenue|sales|profit) so a freshly seeded file renders a populated
# dashboard on first paint.
SAMPLE_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "sample_data")
SAMPLE_TEMPLATES = {
    "uk_smb_sales": {
        "filename": "uk_smb_sales.csv",
        "display_name": "Sample — UK SMB Sales Performance.csv",
    },
    "marketing_campaigns": {
        "filename": "marketing_campaigns.csv",
        "display_name": "Sample — Marketing Campaigns.csv",
    },
    "operations_pipeline": {
        "filename": "operations_pipeline.csv",
        "display_name": "Sample — Operations Pipeline.csv",
    },
}


# ── R2 / S3 helpers ──────────────────────────────────────────────────────────

def _get_s3_client():
    import boto3
    return boto3.client(
        "s3",
        region_name=settings.AWS_REGION,
        endpoint_url=settings.AWS_ENDPOINT_URL,
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
    )

def save_file_r2(file_bytes: bytes, filename: str) -> str:
    unique_name = str(uuid.uuid4()) + "_" + filename
    s3 = _get_s3_client()
    s3.put_object(
        Bucket=settings.AWS_BUCKET_NAME,
        Key=unique_name,
        Body=file_bytes,
    )
    return unique_name

def get_file_r2(s3_key: str) -> bytes:
    s3 = _get_s3_client()
    response = s3.get_object(Bucket=settings.AWS_BUCKET_NAME, Key=s3_key)
    return response["Body"].read()

def delete_file_r2(s3_key: str):
    s3 = _get_s3_client()
    s3.delete_object(Bucket=settings.AWS_BUCKET_NAME, Key=s3_key)


# ── Local storage helper ──────────────────────────────────────────────────────

def save_file_local(file_bytes: bytes, filename: str) -> str:
    os.makedirs(settings.LOCAL_UPLOAD_DIR, exist_ok=True)
    unique_name = str(uuid.uuid4()) + "_" + filename
    path = os.path.join(settings.LOCAL_UPLOAD_DIR, unique_name)
    with open(path, "wb") as f:
        f.write(file_bytes)
    return unique_name


# ── Quota check ───────────────────────────────────────────────────────────────

def check_upload_quota(org: Organisation, db: Session) -> bool:
    from sqlalchemy import func
    from database import DataFile
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    count = db.query(DataFile).filter(
        DataFile.organisation_id == org.id,
        DataFile.created_at >= month_start
    ).count()
    return count < org.max_uploads_per_month


# ── Ingestion helper ─────────────────────────────────────────────────────────
# Shared persistence path used by both /upload (real user file) and
# /seed-sample (bundled demo file). Keeping them funnelled through one helper
# means the DataFile row, storage key, and parsed metadata always look
# identical regardless of ingestion source — every downstream analytics
# endpoint treats a seeded sample exactly like a real upload.

def _ingest_bytes(
    content: bytes,
    filename: str,
    org: Organisation,
    current_user: User,
    db: Session,
):
    """Persist a file and parse headers + a small preview.

    Returns (db_file, meta) where meta carries warnings and a 5-row preview that
    the frontend can surface to the user immediately after upload.
    """
    ext = os.path.splitext(filename)[1].lower()

    # Storage — R2 if configured, else local disk
    use_r2 = settings.STORAGE_TYPE == "s3" and settings.AWS_BUCKET_NAME
    if use_r2:
        storage_key = save_file_r2(content, filename)
        storage_type = "s3"
        file_content_db = None
    else:
        storage_key = save_file_local(content, filename)
        storage_type = "local"
        file_content_db = content

    # Parse headers + row count + preview. Capture warnings so the UI can
    # surface silent failures (encoding drops, mismatched column counts).
    row_count = None
    col_count = None
    columns = []
    preview_rows = []
    warnings = []
    try:
        import io
        if ext in [".xlsx", ".xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1), None) or []
            raw_headers = [cell.value for cell in header_row]
            columns = [str(h) for h in raw_headers if h is not None]
            row_count = max(0, (ws.max_row or 1) - 1)
            col_count = len(columns)
            # Pull up to 5 data rows for preview
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
                if i >= 5:
                    break
                preview_rows.append({
                    str(h): ("" if v is None else v)
                    for h, v in zip(columns, row)
                })
        elif ext == ".csv":
            import csv
            # Try strict decode first. If it fails, fall back to replacement
            # and count the bytes that couldn't decode as a warning signal.
            text = None
            lossy = False
            for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
                try:
                    text = content.decode(enc)
                    if enc != "utf-8":
                        warnings.append(f"File was decoded as {enc} (not UTF-8). Some characters may display differently.")
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                text = content.decode("utf-8", errors="replace")
                lossy = True
                # Count replacement characters as a rough encoding-loss metric
                repl_count = text.count("\ufffd")
                warnings.append(
                    f"File had {repl_count:,} characters that couldn't be decoded and were replaced. "
                    "Please re-save the file as UTF-8 for best results."
                )

            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows:
                columns = rows[0]
                row_count = len(rows) - 1
                col_count = len(columns)
                # Detect rows that don't match the column count — a common
                # symptom of quoting / delimiter issues that would otherwise
                # fail silently later.
                bad_row_count = sum(1 for r in rows[1:] if len(r) != col_count)
                if bad_row_count:
                    warnings.append(
                        f"{bad_row_count:,} row(s) did not match the header column count. "
                        "This can indicate unquoted commas or line-breaks in values."
                    )
                for r in rows[1:6]:
                    preview_rows.append({
                        columns[i] if i < len(columns) else f"col_{i}": v
                        for i, v in enumerate(r)
                    })
    except Exception as e:
        warnings.append(f"Could not fully parse file: {e}")

    # Content hash for duplicate detection (cheap: just SHA-256 of the bytes).
    content_hash = hashlib.sha256(content).hexdigest()

    db_file = DataFile(
        id=str(uuid.uuid4()),
        filename=storage_key,
        original_filename=filename,
        file_size=len(content),
        row_count=row_count,
        column_count=col_count,
        columns_json=json.dumps(columns),
        s3_key=storage_key,
        storage_type=storage_type,
        file_content=file_content_db,
        organisation_id=org.id,
        uploaded_by=current_user.id,
    )
    # Store the hash on the model if the column exists (forward-compatible
    # with a future migration). We don't rely on it for dedup — that path uses
    # an on-the-fly hash against prior content when needed.
    if hasattr(db_file, "content_hash"):
        db_file.content_hash = content_hash

    db.add(db_file)
    db.commit()
    db.refresh(db_file)

    meta = {
        "warnings": warnings,
        "preview_rows": preview_rows,
        "content_hash": content_hash,
    }
    return db_file, meta


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, and .csv files are allowed")

    org = current_user.organisation
    if not org:
        raise HTTPException(status_code=403, detail="No organisation found")

    if not check_upload_quota(org, db):
        raise HTTPException(status_code=429, detail="Monthly upload limit reached. Please upgrade your plan.")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 500MB.")

    db_file, meta = _ingest_bytes(content, file.filename, org, current_user, db)

    return {
        "id": db_file.id,
        "filename": db_file.original_filename,
        "size": db_file.file_size,
        "rows": db_file.row_count,
        "columns": db_file.column_count,
        "column_names": json.loads(db_file.columns_json) if db_file.columns_json else [],
        "uploaded_at": db_file.created_at.isoformat(),
        "warnings": meta.get("warnings") or [],
        "preview_rows": meta.get("preview_rows") or [],
        "content_hash": meta.get("content_hash"),
    }


@router.post("/seed-sample/{template_id}")
def seed_sample(
    template_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Copy a bundled demo CSV into the caller's org so a brand-new user can
    see a populated Executive Dashboard in <30 seconds without uploading
    anything. Purely additive — existing files are not touched. The copy
    flows through the same _ingest_bytes path as a real upload, so every
    dashboard treats it identically."""
    template = SAMPLE_TEMPLATES.get(template_id)
    if not template:
        raise HTTPException(status_code=404, detail=f"Unknown sample template: {template_id}")

    org = current_user.organisation
    if not org:
        raise HTTPException(status_code=403, detail="No organisation found")

    # Skip quota check intentionally — seeding a sample is an onboarding
    # action and shouldn't be gated by upload limits. Users who hit their
    # real upload cap still get this one-click "see a demo" escape hatch.

    src_path = os.path.join(SAMPLE_DATA_DIR, template["filename"])
    if not os.path.exists(src_path):
        raise HTTPException(status_code=500, detail=f"Sample file missing on server: {template['filename']}")

    with open(src_path, "rb") as fh:
        content = fh.read()

    db_file, _meta = _ingest_bytes(content, template["display_name"], org, current_user, db)

    return {
        "id": db_file.id,
        "filename": db_file.original_filename,
        "size": db_file.file_size,
        "rows": db_file.row_count,
        "columns": db_file.column_count,
        "column_names": json.loads(db_file.columns_json) if db_file.columns_json else [],
        "uploaded_at": db_file.created_at.isoformat(),
        "is_sample": True,
        "template_id": template_id,
    }


class DuplicateCheck(BaseModel):
    filename: str
    size: int


@router.post("/check-duplicate")
def check_duplicate(
    body: DuplicateCheck,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Called from the Files page BEFORE a user uploads a big file. If another
    file in this org has exactly the same name AND byte size, return it so the
    UI can offer "use existing" instead of re-ingesting the same data. Cheap
    approximation of a real content-hash dedup; doesn't require the user to
    wait for a full upload to find out they duplicated something.
    """
    if not current_user.organisation_id:
        return {"match": None}
    match = (
        db.query(DataFile)
        .filter(
            DataFile.organisation_id == current_user.organisation_id,
            DataFile.original_filename == body.filename,
            DataFile.file_size == body.size,
        )
        .order_by(DataFile.created_at.desc())
        .first()
    )
    if not match:
        return {"match": None}
    return {
        "match": {
            "id": match.id,
            "filename": match.original_filename,
            "size": match.file_size,
            "row_count": match.row_count,
            "column_count": match.column_count,
            "uploaded_at": match.created_at.isoformat() if match.created_at else None,
        }
    }


@router.get("/sample-templates")
def list_sample_templates(
    current_user: User = Depends(get_current_user),
):
    """Expose the catalog so the frontend can render the right options
    on the empty-state card without hardcoding template IDs."""
    return [
        {
            "id": tid,
            "display_name": t["display_name"],
            "filename": t["filename"],
        }
        for tid, t in SAMPLE_TEMPLATES.items()
    ]


@router.get("/")
def list_files(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    files = db.query(DataFile).filter(
        DataFile.organisation_id == current_user.organisation_id
    ).order_by(DataFile.created_at.desc()).all()

    return [
        {
            "id": f.id,
            "filename": f.original_filename,
            "size": f.file_size,
            "row_count": f.row_count,
            "column_count": f.column_count,
            "column_names": json.loads(f.columns_json) if f.columns_json else [],
            "storage_type": f.storage_type,
            "last_synced_at": getattr(f, "last_synced_at", None) and getattr(f, "last_synced_at").isoformat(),
            "uploaded_at": f.created_at.isoformat()
        }
        for f in files
    ]


@router.get("/{file_id}/download")
def get_file_data(
    file_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    f = db.query(DataFile).filter(
        DataFile.id == file_id,
        DataFile.organisation_id == current_user.organisation_id
    ).first()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")

    import base64

    if f.storage_type == "s3":
        try:
            data_bytes = get_file_r2(f.s3_key)
        except Exception as e:
            raise HTTPException(status_code=404, detail=f"File not found in R2: {str(e)}")
        data = base64.b64encode(data_bytes).decode()
    else:
        path = os.path.join(settings.LOCAL_UPLOAD_DIR, f.filename)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found on disk")
        with open(path, "rb") as fp:
            data = base64.b64encode(fp.read()).decode()

    return {"filename": f.original_filename, "data": data, "encoding": "base64"}


@router.delete("/{file_id}")
def delete_file(
    file_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    f = db.query(DataFile).filter(
        DataFile.id == file_id,
        DataFile.organisation_id == current_user.organisation_id
    ).first()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")

    if f.storage_type == "s3":
        try:
            delete_file_r2(f.s3_key)
        except Exception:
            pass  # Best-effort: still remove DB record
    else:
        path = os.path.join(settings.LOCAL_UPLOAD_DIR, f.filename)
        if os.path.exists(path):
            os.remove(path)

    db.delete(f)
    db.commit()
    return {"message": "File deleted"}
